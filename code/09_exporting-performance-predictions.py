import pandas as pd
from openpyxl import load_workbook
import os


#-----------------------------------------------------------------------------------------------
#LOAD IN UNCLEANED ORGINAL DATA AND SPECIFY LOCATION OF EXPORTING
#-----------------------------------------------------------------------------------------------

# Load in the original data before cleaning
original_df = pd.read_csv(r'data/INSERT ORIGINAL DATA PATH HERE PRIOR TO CLEANING.csv')

# Filter rows where our DataFrame includes only specified values in the target column
unlabeled_original_df = original_df[
    original_df['Targert_col'].isin(["Var1", "Var2"])
]

# Drop columns that are not needed for our predictions sheet
drop_columns = ['COLUMN_NAME_1', 'COLUMN_NAME_2']  
unlabeled_trimmed_df = unlabeled_original_df.drop(columns=drop_columns)

# Prepare Excel file for logging results specify the sheet name you want and file you need regardless if it exists
model_eval_path = r'data/model_eval/model_evaluations.xlsx'
sheet_name_performance = 'Model Performance'
sheet_name_predictions = 'Model Predictions'


#-----------------------------------------------------------------------------------------------
#BELOW IS A FUNCTION TO EXPORT OUR MODEL PERFORMANCE METRICS 
#-----------------------------------------------------------------------------------------------

def export_performance(model_eval_path, sheet_name_performance, model, stop_words_removed, 
                       input_data, input_type, training_size_dist, precision_results, final_pred_dist):
    """
    Export model performance metrics to an Excel file.
    
    Parameters:
    - model_eval_path (str): Path to the Excel file for storing performance metrics.
    - sheet_name_performance (str): Name of the sheet for performance metrics.
    - model (object): The model object (e.g., RandomForestClassifier).
    - stop_words_removed (bool): Whether stop words were removed.
    - input_data (str): Description of input data (e.g., "all text", "abstract").
    - input_type (str): Type of input vectorization/embedding.
    - training_size_dist (str): Pre-formatted string of class distribution.
    - precision_results (dict): Precision at different thresholds, e.g., {'Precision_50': 0.85}.
    - final_pred_dist (str): Final prediction distribution for each class.
    
    Returns:
    None
    """
    # Prepare row data dynamically
    row_data = {
        'Model': str(model),
        'Stop Words Removed': 'Yes' if stop_words_removed else 'No',
        'Input Data': input_data,
        'Input Type': input_type,
        'Training Size Distribution': training_size_dist,
        'Final Prediction Distribution': final_pred_dist
    }
    
    # Add precision metrics dynamically
    row_data.update(precision_results)
    
    # Create DataFrame for the row
    df = pd.DataFrame([row_data])
    
    # Append or create new Excel file
    if os.path.exists(model_eval_path):
        with pd.ExcelWriter(model_eval_path, mode='a', engine='openpyxl') as writer:
            book = load_workbook(model_eval_path)
            writer.book = book

            if sheet_name_performance in writer.book.sheetnames:
                # Append without overwriting existing headers
                df.to_excel(writer, sheet_name=sheet_name_performance, index=False, header=False, 
                            startrow=writer.book[sheet_name_performance].max_row)
            else:
                # Write to a new sheet
                df.to_excel(writer, sheet_name=sheet_name_performance, index=False)
    else:
        # Create a new file and write data
        df.to_excel(model_eval_path, sheet_name=sheet_name_performance, index=False)
    
    print(f"Performance metrics saved to {model_eval_path}, sheet: {sheet_name_performance}")


#-----------------------------------------------------------------------------------------------
#BELOW IS A FUNCTION TO EXPORT OUR MODEL PREDICTIONS ONTO UNLABELED DATA, EXPORT PREDICTIONS OF BEST PERFORMING MODEL
#-----------------------------------------------------------------------------------------------

def export_predictions(model_eval_path, sheet_name_predictions, model_name, predicted_labels, 
                       prediction_probabilities, unlabeled_trimmed_df ):
    """
    Export model predictions and probabilities to an Excel file.
    
    Parameters:
    - model_eval_path (str): Path to the Excel file to load or create.
    - sheet_name_predictions (str): Name of the Excel sheet to read and write data.
    - model_name (str): Name of the model ( ADD THE NAME YOU WANT IT TO BE).
    - predicted_labels (list or Series): Predicted labels from the model.
    - prediction_probabilities (list or Series): Prediction probabilities from the model.
    - unlabeled_trimmed_df (DataFrame): Original DataFrame to use if the file does not exist.
    - model_eval_path (str): Path to save the updated Excel file.
    
    Returns:
    None
    """
    # Load existing Excel file or initialize DataFrame from original unlabeled data
    try:
        with pd.ExcelFile(model_eval_path) as reader:
            predictions_df = pd.read_excel(reader, sheet_name=sheet_name_predictions)
    except FileNotFoundError:
        predictions_df = unlabeled_trimmed_df.copy()

    # Add new model predictions and probabilities
    predictions_df[f"{model_name}_Label"] = predicted_labels
    predictions_df[f"{model_name}_Prob"] = prediction_probabilities

    # Save updated data to Excel
    with pd.ExcelWriter(model_eval_path, engine='openpyxl', mode='a') as writer:
        predictions_df.to_excel(writer, index=False, sheet_name=sheet_name_predictions)

    print(f"Added predicted labels and probabilities for {model_name} to {model_eval_path}")

